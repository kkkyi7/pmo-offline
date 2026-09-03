import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

def build_automated_gantt_excel():
    source_path = r'C:\Users\hf\Downloads\传化项目实施主计划&WBS_20260903.2.xlsx'
    if not os.path.exists(source_path):
        source_path = r'C:\Users\hf\Projects\pmo-offline\templates\项目实施主计划&WBS-V0.7.xlsx'
    
    print(f"Loading source template: {source_path}")
    wb = openpyxl.load_workbook(source_path, data_only=False)
    
    ws = wb['主计划']
    
    # 1. Clear any old empty columns from 21 to ws.max_column
    if ws.max_column >= 21:
        for r in range(1, ws.max_row + 1):
            for c in range(21, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
    
    # Set left columns width & alignment (Cols A to T)
    col_widths = {
        'A': 5.5,   # 一级流程（阶段）
        'B': 6.5,   # 二级流程（场景）
        'C': 8.0,   # 三级流程（任务）
        'D': 30.0,  # 任务名称
        'E': 6.0,   # 后置
        'F': 8.0,   # 前置
        'G': 9.0,   # 进度状态
        'H': 12.0,  # 进度备注
        'I': 7.5,   # 里程碑
        'J': 8.0,   # 乙方评审
        'K': 8.0,   # 甲方评审
        'L': 7.5,   # PMO
        'M': 8.5,   # 工作周
        'N': 13.0,  # 计划开始
        'O': 13.0,  # 计划结束
        'P': 9.0,   # 计划天数
        'Q': 11.0,  # 是否前后置
        'R': 9.5,   # 执行人主责
        'S': 11.0,  # 协助或支持
        'T': 24.0,  # 关键交付物
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Ensure Col N and O date formatting
    for r in range(3, ws.max_row + 1):
        ws.cell(r, 14).number_format = 'yyyy-mm-dd'
        ws.cell(r, 15).number_format = 'yyyy-mm-dd'
        ws.cell(r, 16).alignment = Alignment(horizontal='center', vertical='center')
    
    # 2. Date Range: 2026-08-17 (Monday W0) to 2027-12-31
    start_date = datetime.date(2026, 8, 17)
    end_date = datetime.date(2027, 12, 31)
    total_days = (end_date - start_date).days + 1
    
    start_col = 21 # Col U
    end_col = start_col + total_days - 1
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    
    print(f"Generating Gantt timeline: {start_date} to {end_date} ({total_days} days, Col {start_col_letter} to {end_col_letter})")
    
    # Styles
    month_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # Navy
    month_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    
    day_font_weekday = Font(name='Segoe UI', size=9, bold=False, color='334155')
    day_font_weekend = Font(name='Segoe UI', size=9, bold=True, color='94A3B8')
    day_fill_weekday = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    day_fill_weekend = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
    
    thin_border_side = Side(border_style='thin', color='CBD5E1')
    day_border = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side, right=thin_border_side)
    
    # Map months to their column ranges
    month_ranges = []
    current_month_key = None
    m_start_col = start_col
    
    # Write Row 1 & Row 2
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    
    for i in range(total_days):
        col = start_col + i
        d = start_date + datetime.timedelta(days=i)
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 2.85
        
        # Check month change
        m_key = (d.year, d.month)
        if m_key != current_month_key:
            if current_month_key is not None:
                month_ranges.append((current_month_key, m_start_col, col - 1))
            current_month_key = m_key
            m_start_col = col
        
        # Row 2: Date
        day_cell = ws.cell(2, col, d)
        day_cell.number_format = 'd'
        is_weekend = d.weekday() >= 5
        day_cell.font = day_font_weekend if is_weekend else day_font_weekday
        day_cell.fill = day_fill_weekend if is_weekend else day_fill_weekday
        day_cell.alignment = Alignment(horizontal='center', vertical='center')
        day_cell.border = day_border
    
    # Add last month
    if current_month_key is not None:
        month_ranges.append((current_month_key, m_start_col, end_col))
    
    # Merge and style month headers in Row 1
    for (year, month), c1, c2 in month_ranges:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        m_cell = ws.cell(1, c1, f"{year}年{month}月")
        m_cell.fill = month_fill
        m_cell.font = month_font
        m_cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(c1, c2 + 1):
            ws.cell(1, c).fill = month_fill
            ws.cell(1, c).border = day_border
    
    # 3. Outline Grouping (数据分级显示 折叠展开)
    max_task_row = ws.max_row
    while max_task_row > 3 and not ws.cell(max_task_row, 4).value:
        max_task_row -= 1
        
    print(f"Task rows: 3 to {max_task_row}")
    
    phase_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    phase_font = Font(name='微软雅黑', size=10, bold=True, color='0F172A')
    scene_font = Font(name='微软雅黑', size=9.5, bold=True, color='0F766E')
    
    for r in range(3, max_task_row + 1):
        ws.row_dimensions[r].height = 20
        l1 = ws.cell(r, 1).value
        l2 = ws.cell(r, 2).value
        l3 = ws.cell(r, 3).value
        
        if l1 is not None and str(l1).strip():
            # Phase Level 1
            ws.row_dimensions[r].outline_level = 0
            for c in range(1, 21):
                cell = ws.cell(r, c)
                cell.fill = phase_fill
                cell.font = phase_font
        elif l2 is not None and str(l2).strip():
            # Scene Level 2
            ws.row_dimensions[r].outline_level = 1
            ws.cell(r, 4).font = scene_font
        elif l3 is not None and str(l3).strip():
            # Task Level 3
            ws.row_dimensions[r].outline_level = 2
        else:
            ws.row_dimensions[r].outline_level = 2

    # 4. Freeze Panes: Lock top 2 rows and left 20 columns
    ws.freeze_panes = f"{start_col_letter}3"
    
    # 5. Conditional Formatting Rules (Auto Gantt Bars)
    gantt_range = f"{start_col_letter}3:{end_col_letter}{max_task_row}"
    print(f"Adding conditional formatting rules to range: {gantt_range}")
    
    # Colors
    fill_phase = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')      # Dark navy
    fill_scene = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')      # Teal
    fill_task = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')        # Sky blue
    fill_done = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')        # Emerald green
    fill_milestone = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')   # Amber gold
    fill_weekend = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')     # Light gray weekend
    
    # Priority 1: Milestone
    rule_milestone = FormulaRule(
        formula=[f'AND({start_col_letter}$2=$N3, $I3<>"")'],
        fill=fill_milestone
    )
    
    # Priority 2: Completed tasks
    rule_done = FormulaRule(
        formula=[f'AND({start_col_letter}$2>=$N3, {start_col_letter}$2<=$O3, $G3="已完成")'],
        fill=fill_done
    )

    # Priority 3: Phase level 1 bar
    rule_phase = FormulaRule(
        formula=[f'AND({start_col_letter}$2>=$N3, {start_col_letter}$2<=$O3, $A3<>"")'],
        fill=fill_phase
    )

    # Priority 4: Scene level 2 bar
    rule_scene = FormulaRule(
        formula=[f'AND({start_col_letter}$2>=$N3, {start_col_letter}$2<=$O3, $B3<>"")'],
        fill=fill_scene
    )

    # Priority 5: Regular task bar
    rule_task = FormulaRule(
        formula=[f'AND({start_col_letter}$2>=$N3, {start_col_letter}$2<=$O3)'],
        fill=fill_task
    )

    # Priority 6: Weekend background
    rule_weekend = FormulaRule(
        formula=[f'WEEKDAY({start_col_letter}$2, 2)>5'],
        fill=fill_weekend
    )

    ws.conditional_formatting.add(gantt_range, rule_milestone)
    ws.conditional_formatting.add(gantt_range, rule_done)
    ws.conditional_formatting.add(gantt_range, rule_phase)
    ws.conditional_formatting.add(gantt_range, rule_scene)
    ws.conditional_formatting.add(gantt_range, rule_task)
    ws.conditional_formatting.add(gantt_range, rule_weekend)

    # 6. Enable full auto calculation on load
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    
    # Save targets
    desktop_file_ch = r'C:\Users\hf\Desktop\传化项目实施主计划&WBS_2026-2027_自动甘特版.xlsx'
    desktop_file_tpl = r'C:\Users\hf\Desktop\项目实施主计划&WBS-2026-2027_自动甘特版.xlsx'
    repo_file = r'C:\Users\hf\Projects\pmo-offline\templates\项目实施主计划&WBS-2026-2027_自动甘特版.xlsx'
    
    wb.save(desktop_file_ch)
    print(f"Saved: {desktop_file_ch}")
    wb.save(desktop_file_tpl)
    print(f"Saved: {desktop_file_tpl}")
    wb.save(repo_file)
    print(f"Saved: {repo_file}")
    print("ALL DONE!")

if __name__ == '__main__':
    build_automated_gantt_excel()
